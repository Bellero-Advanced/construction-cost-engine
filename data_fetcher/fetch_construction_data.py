#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 Construction Material Price Data Fetcher
 ดึงข้อมูลราคาวัสดุก่อสร้างจาก API จริงของหน่วยงานภาครัฐไทย
================================================================================

สคริปต์นี้ดึงข้อมูล "ของจริง" (ไม่ใช่ mock) จาก API สาธารณะ 3 แหล่ง:

  1. MOC CSI  (dataapi.moc.go.th/csi-indexes)
     → ดัชนีราคาวัสดุก่อสร้างระดับประเทศ รายเดือน
        Construction Sector Index จากกระทรวงพาณิชย์

  2. MOC CPIP (dataapi.moc.go.th/cpip-indexes)
     → ดัชนีราคาก่อสร้างรายจังหวัด (77 จังหวัด) รายเดือน
        ใช้สำหรับราคากลางงานก่อสร้างภาครัฐ

  3. Green Label (nhicservices.nha.co.th/.../green_label/mid)
     → ราคาวัสดุก่อสร้างฉลากเขียวจริง หน่วยเป็นบาท (5,000+ รายการ)
        มียี่ห้อ สเปค หน่วย และราคาต่อหน่วย

ทุกครั้งที่รัน:
  - สร้างโฟลเดอร์ output/ ให้อัตโนมัติ
  - ได้ไฟล์ Excel ใหม่ 1 ไฟล์ (ตั้งชื่อตาม timestamp กันชนกัน)
  - บันทึก log การดึงทุกขั้นตอนลงใน sheet "Fetch_Logs"

วิธีรัน:
    python3 fetch_construction_data.py

Dependencies:
    pip install pandas openpyxl

--------------------------------------------------------------------------------
Author : Construction Cost Engine Project
Updated: 2026-06-10
--------------------------------------------------------------------------------
"""

import urllib.request
import urllib.error
import json
import datetime
import os
import sys
import time
from typing import Dict, List, Optional, Any

# ----- ตรวจสอบ dependencies ก่อนเริ่ม -----
try:
    import pandas as pd
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    print(f"[X] ขาด dependency: {exc}")
    print("    ติดตั้งด้วย: pip install pandas openpyxl")
    sys.exit(1)


# ==============================================================================
#  CONFIG — ปรับค่าได้ที่นี่
# ==============================================================================
class Config:
    OUTPUT_DIR = "output"          # โฟลเดอร์เก็บไฟล์ Excel
    TIMEOUT = 45                   # วินาที ต่อ 1 request (API ภาครัฐช้า)
    RETRIES = 2                    # จำนวนครั้งที่ลองใหม่ถ้า fail
    RETRY_DELAY = 3                # หน่วงกี่วินาทีก่อน retry
    PACE_DELAY = 1.0               # หน่วงระหว่าง request กัน rate-limit

    # ช่วงปีที่ต้องการดึง (ค.ศ.)
    CSI_FROM_YEAR = 2023
    CSI_TO_YEAR = 2025
    CPIP_FROM_YEAR = 2024
    CPIP_TO_YEAR = 2025

    USER_AGENT = "Mozilla/5.0 (ConstructionCostEngine/1.0; +data-fetcher)"


# ==============================================================================
#  HTTP helper
# ==============================================================================
def http_get_json(url: str, timeout: int, retries: int,
                  retry_delay: int, logger=print) -> Optional[Any]:
    """
    ยิง GET request แล้ว parse JSON พร้อม retry

    คืนค่า: dict/list ที่ parse แล้ว หรือ None ถ้า fail ทุกครั้ง
    """
    last_err = None
    for attempt in range(retries + 1):
        try:
            req = urllib.request.Request(
                url,
                headers={
                    "User-Agent": Config.USER_AGENT,
                    "Accept": "application/json",
                },
            )
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
                return json.loads(raw)
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as err:
            last_err = err
            if attempt < retries:
                logger(f"      retry {attempt + 1}/{retries} "
                       f"({type(err).__name__}) ...")
                time.sleep(retry_delay)
            else:
                logger(f"      [X] ล้มเหลวหลังลอง {retries + 1} ครั้ง: "
                       f"{type(err).__name__}: {err}")
    return None


# ==============================================================================
#  Main fetcher class
# ==============================================================================
class ConstructionDataFetcher:
    """ดึงข้อมูลราคาวัสดุก่อสร้างจาก API ภาครัฐ แล้วเขียนลง Excel"""

    MOC_BASE = "https://dataapi.moc.go.th"
    GREEN_LABEL_URL = "https://nhicservices.nha.co.th/common/services/green_label/mid"

    def __init__(self, output_dir: str = Config.OUTPUT_DIR):
        self.output_dir = output_dir
        self.logs: List[str] = []
        self.run_started = datetime.datetime.now()
        # สร้างโฟลเดอร์เก็บผลลัพธ์ (ถ้ายังไม่มี)
        os.makedirs(output_dir, exist_ok=True)
        self.log(f"โฟลเดอร์ output: {os.path.abspath(output_dir)}")

    # ---------- logging ----------
    def log(self, message: str):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {message}"
        print(line)
        self.logs.append(line)

    # ---------- SOURCE 1: CSI national index ----------
    def fetch_csi_national(self) -> Optional[pd.DataFrame]:
        """
        ดึงดัชนีราคาวัสดุก่อสร้างระดับประเทศ (CSI)
        index_id 0000000000000000 = ดัชนีรวม
        """
        self.log("")
        self.log("[1/3] CSI — ดัชนีราคาวัสดุก่อสร้างระดับประเทศ (รายเดือน)")
        url = (f"{self.MOC_BASE}/csi-indexes"
               f"?index_id=0000000000000000"
               f"&from_year={Config.CSI_FROM_YEAR}"
               f"&to_year={Config.CSI_TO_YEAR}")
        self.log(f"      GET {url}")

        data = http_get_json(url, Config.TIMEOUT, Config.RETRIES,
                             Config.RETRY_DELAY, self.log)
        if not data or not isinstance(data, list):
            self.log("      [X] ไม่ได้ข้อมูล CSI")
            return None

        df = pd.DataFrame(data)
        df = self._add_thai_date(df)
        self.log(f"      [OK] {len(df)} แถว "
                 f"({df['year'].min()}/{df['month'].min()} - "
                 f"{df['year'].max()}/{df['month'].max()})")
        time.sleep(Config.PACE_DELAY)
        return df

    # ---------- SOURCE 2: CPIP provincial index ----------
    def fetch_cpip_provincial(self) -> Optional[pd.DataFrame]:
        """
        ดึงดัชนีราคาก่อสร้างรายจังหวัด (CPIP) — 77 จังหวัด
        ใช้สำหรับราคากลางงานก่อสร้างภาครัฐ
        """
        self.log("")
        self.log("[2/3] CPIP — ดัชนีราคาก่อสร้างรายจังหวัด (77 จังหวัด)")
        url = (f"{self.MOC_BASE}/cpip-indexes"
               f"?index_id=0000000000000000"
               f"&from_year={Config.CPIP_FROM_YEAR}"
               f"&to_year={Config.CPIP_TO_YEAR}")
        self.log(f"      GET {url}")

        data = http_get_json(url, Config.TIMEOUT, Config.RETRIES,
                             Config.RETRY_DELAY, self.log)
        if not data or not isinstance(data, list):
            self.log("      [X] ไม่ได้ข้อมูล CPIP")
            return None

        df = pd.DataFrame(data)
        df = self._add_thai_date(df)
        n_prov = df["province_name"].nunique() if "province_name" in df else 0
        self.log(f"      [OK] {len(df)} แถว, {n_prov} จังหวัด")
        time.sleep(Config.PACE_DELAY)
        return df

    # ---------- SOURCE 3: Green Label real prices ----------
    def fetch_green_label(self) -> Optional[pd.DataFrame]:
        """
        ดึงราคาวัสดุก่อสร้างฉลากเขียว — ราคาจริงเป็นบาท
        โครงสร้าง response: {success, message, result:{total_rows, items:[...]}}
        """
        self.log("")
        self.log("[3/3] Green Label — ราคาวัสดุฉลากเขียวจริง (บาท)")
        self.log(f"      GET {self.GREEN_LABEL_URL}")

        data = http_get_json(self.GREEN_LABEL_URL, Config.TIMEOUT,
                            Config.RETRIES, Config.RETRY_DELAY, self.log)
        if not data or not data.get("success"):
            self.log("      [X] ไม่ได้ข้อมูล Green Label")
            return None

        result = data.get("result", {})
        items = result.get("items", []) if isinstance(result, dict) else []
        if not items:
            self.log("      [X] result ว่าง")
            return None

        df = pd.DataFrame(items)
        # แปลง price_baht "2,731.10" → float
        if "price_baht" in df.columns:
            df["price_baht_numeric"] = (
                df["price_baht"].astype(str)
                .str.replace(",", "", regex=False)
                .apply(lambda x: pd.to_numeric(x, errors="coerce"))
            )
        n_cat = df["category"].nunique() if "category" in df else 0
        self.log(f"      [OK] {len(df)} รายการ, {n_cat} หมวดหมู่")
        time.sleep(Config.PACE_DELAY)
        return df

    # ---------- helper: เพิ่มคอลัมน์วันที่ไทย ----------
    @staticmethod
    def _add_thai_date(df: pd.DataFrame) -> pd.DataFrame:
        """เพิ่มคอลัมน์ period (YYYY-MM) จาก year/month ถ้ามี"""
        if "year" in df.columns and "month" in df.columns:
            df = df.copy()
            df.insert(
                0, "period",
                df["year"].astype(str) + "-" +
                df["month"].astype(str).str.zfill(2)
            )
        return df

    # ---------- สร้าง Green Label summary ----------
    @staticmethod
    def build_green_summary(green_df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """สรุปราคาเฉลี่ย/ต่ำสุด/สูงสุด ต่อหมวดหมู่"""
        if green_df is None or "price_baht_numeric" not in green_df.columns:
            return None
        grp = green_df.dropna(subset=["price_baht_numeric"]).groupby("category")
        summary = grp["price_baht_numeric"].agg(
            count="count", min_price="min",
            avg_price="mean", max_price="max"
        ).reset_index()
        summary = summary.sort_values("count", ascending=False)
        summary["avg_price"] = summary["avg_price"].round(2)
        return summary

    # ==========================================================================
    #  EXCEL WRITER
    # ==========================================================================
    def write_excel(self, frames: Dict[str, pd.DataFrame]) -> str:
        """
        เขียนทุก DataFrame ลง Excel ไฟล์เดียว หลาย sheet
        แล้วจัด format หัวตาราง + ความกว้างคอลัมน์
        """
        ts = self.run_started.strftime("%Y%m%d_%H%M%S")
        filename = f"construction_data_{ts}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        self.log("")
        self.log(f"เขียน Excel: {filename}")

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # --- Sheet: Summary (ภาพรวมการดึง) ---
            self._write_summary_sheet(writer, frames)

            # --- Data sheets ---
            sheet_specs = [
                ("CSI_National", "csi", "ดัชนีราคาวัสดุก่อสร้างระดับประเทศ"),
                ("CPIP_Provincial", "cpip", "ดัชนีราคาก่อสร้างรายจังหวัด"),
                ("GreenLabel_Prices", "green", "ราคาวัสดุฉลากเขียว (บาท)"),
                ("GreenLabel_Summary", "green_summary", "สรุปราคาตามหมวดหมู่"),
            ]
            for sheet_name, key, _desc in sheet_specs:
                df = frames.get(key)
                if df is not None and len(df) > 0:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self.log(f"      sheet: {sheet_name} ({len(df)} แถว)")

            # --- Sheet: Fetch_Logs ---
            logs_df = pd.DataFrame({"timestamp_log": self.logs})
            logs_df.to_excel(writer, sheet_name="Fetch_Logs", index=False)
            self.log(f"      sheet: Fetch_Logs ({len(self.logs)} บรรทัด)")

        # จัด format หลังเขียนเสร็จ
        self._style_workbook(filepath)
        self.log(f"[DONE] บันทึกไฟล์: {os.path.abspath(filepath)}")
        return filepath

    def _write_summary_sheet(self, writer, frames: Dict[str, pd.DataFrame]):
        """sheet แรก = ภาพรวมการดึงข้อมูล"""
        def n(key):
            df = frames.get(key)
            return len(df) if df is not None else 0

        rows = [
            ("เวลาที่ดึงข้อมูล", self.run_started.strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("แหล่งข้อมูล (Source)", "API"),
            ("1. CSI ดัชนีราคาวัสดุระดับประเทศ", f"{n('csi')} แถว"),
            ("2. CPIP ดัชนีราคารายจังหวัด", f"{n('cpip')} แถว"),
            ("3. Green Label ราคาวัสดุจริง", f"{n('green')} รายการ"),
            ("", ""),
            ("CSI endpoint", "dataapi.moc.go.th/csi-indexes"),
            ("CPIP endpoint", "dataapi.moc.go.th/cpip-indexes"),
            ("Green Label endpoint", "nhicservices.nha.co.th/.../green_label/mid"),
            ("", ""),
            ("ช่วงปี CSI", f"{Config.CSI_FROM_YEAR}-{Config.CSI_TO_YEAR}"),
            ("ช่วงปี CPIP", f"{Config.CPIP_FROM_YEAR}-{Config.CPIP_TO_YEAR}"),
        ]
        summary_df = pd.DataFrame(rows, columns=["รายการ", "ค่า"])
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        self.log("      sheet: Summary")

    @staticmethod
    def _style_workbook(filepath: str):
        """จัด format: หัวตารางสีน้ำเงิน + auto column width"""
        from openpyxl import load_workbook
        wb = load_workbook(filepath)

        header_fill = PatternFill("solid", fgColor="1A3556")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            # หัวตาราง (แถวแรก)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            # auto width (จำกัดไม่เกิน 60)
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for c in col_cells:
                    if c.value is not None:
                        max_len = max(max_len, len(str(c.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
            # freeze หัวตาราง
            ws.freeze_panes = "A2"

        wb.save(filepath)


# ==============================================================================
#  MAIN
# ==============================================================================
def main():
    print("=" * 70)
    print("  CONSTRUCTION MATERIAL PRICE DATA FETCHER")
    print("  ดึงข้อมูลราคาวัสดุก่อสร้างจาก API จริงของภาครัฐไทย")
    print("=" * 70)

    fetcher = ConstructionDataFetcher(output_dir=Config.OUTPUT_DIR)

    # ดึงข้อมูล 3 แหล่ง
    csi = fetcher.fetch_csi_national()
    cpip = fetcher.fetch_cpip_provincial()
    green = fetcher.fetch_green_label()
    green_summary = fetcher.build_green_summary(green) if green is not None else None

    frames = {
        "csi": csi,
        "cpip": cpip,
        "green": green,
        "green_summary": green_summary,
    }

    # ตรวจว่าได้ข้อมูลอย่างน้อย 1 แหล่ง
    got_any = any(df is not None and len(df) > 0 for df in frames.values())
    if not got_any:
        fetcher.log("")
        fetcher.log("[X] ไม่ได้ข้อมูลจากแหล่งใดเลย — ตรวจสอบอินเทอร์เน็ต/API")
        # ยังเขียน Excel เพื่อเก็บ log ไว้ debug
    # เขียน Excel เสมอ (มี Fetch_Logs ติดไปด้วยทุกครั้ง)
    filepath = fetcher.write_excel(frames)

    # สรุปท้าย
    print()
    print("=" * 70)
    print("  เสร็จสิ้น")
    print("=" * 70)
    print(f"  ไฟล์ผลลัพธ์ : {filepath}")
    print(f"  CSI         : {len(csi) if csi is not None else 0} แถว")
    print(f"  CPIP        : {len(cpip) if cpip is not None else 0} แถว")
    print(f"  Green Label : {len(green) if green is not None else 0} รายการ")
    print("=" * 70)
    print("  เปิดไฟล์ Excel เพื่อดูข้อมูล — รันสคริปต์ซ้ำได้ทุกเมื่อเพื่อดึงข้อมูลใหม่")
    print("=" * 70)


if __name__ == "__main__":
    main()
